import json

from ryu.base import app_manager
from ryu import cfg
from ryu.app import simple_switch_13
from operator import attrgetter
from ryu.ofproto import ofproto_v1_3
from ryu.controller.handler import set_ev_cls
from ryu.controller import ofp_event
from ryu.controller.handler import MAIN_DISPATCHER, DEAD_DISPATCHER, CONFIG_DISPATCHER, HANDSHAKE_DISPATCHER
from ryu.lib import hub
from ryu.app.wsgi import ControllerBase
from ryu.app.wsgi import Response
from ryu.app.wsgi import route
from ryu.app.wsgi import WSGIApplication

from ryu.lib.packet import packet
from ryu.lib.packet import ethernet

# 用于路径选择
# 导入这些主要是为了让网络链路中产生LLDP数据包，只有产生了LLDP数据报，才能进行LLDP时延探测
from ryu.topology.api import get_switch, get_link, get_host
from ryu.topology import event, switches

# networkx用于存储链路信息，本程序存储为一个有向图
import networkx as nx

import openpyxl
import datetime
import time
#import setting
import copy


monitor_instance_name = 'monitor_api_app'
urlQuery = '/monitor/mactable/getinfo'
urlTotalData = '/monitor/getTotalData'
urlSetDtc = '/monitor/SetDtc/{cid}'

class Monitor(simple_switch_13.SimpleSwitch13):   #继承simple_switch_13的功能
	OFP_VERSIONS = [ofproto_v1_3.OFP_VERSION]	#set openflow version 1.3
	_CONTEXTS = {'wsgi': WSGIApplication}

	def __init__(self, *args, **kwargs):	  #初始化函数
		super(Monitor, self).__init__(*args, **kwargs)
		self.datapaths = {}   #初始化成员变量，用来存储数据
		self.monitor_thread = hub.spawn(self._monitor)   #用协程方法执行_monitor方法，这样其他方法可以被其他协程执行。 hub.spawn()创建协程
		self.monitor_info = {}
		#记录流量差，用于计算链路速率
		self.rx_prebyte = 0
		self.rx_curbyte = 0
		self.tx_prebyte = 0
		self.tx_curbyte = 0
		wsgi = kwargs['wsgi']
		wsgi.register(Monitor_Controller,
						{monitor_instance_name: self})
		#统计总流量信息
		self.workbook = openpyxl.load_workbook("DateInfo.xlsx")
		self.worksheet = self.workbook.get_sheet_by_name('Sheet1')
		self.count = 0
		self.NowDate = str(datetime.date.today())

		#路径选择
		self.choice = 0 #0:无选择， 1:最低延迟， 2:最小跳数
		# 初始化networkx的有向图
		self.G = nx.DiGraph()
		self.topology_api_app = self

		# 存储网络拓扑的交换机id
		self.dpidSwitch = {}
		# 存储echo往返时延
		self.echoDelay = {}
		# 存储LLDP时延
		self.src_dstDelay = {}
		# 存储链路的时延，即LLDP时延-echo的时延，计算出的每条链路的时延
		self.link_Delay = {}

		# 存储源-目的-权重(时延)的列表，用于向有向图写入边信息
		self.links_src_dst = []
		# 存储整个链路各个节点之间的连接信息，包括源端口，
		# 例如s1-s2，通过s1的2端口连接，存储的信息即为：{’1-2‘：2}
		self.id_port = {}
		#用于最短路径
		self.paths = {}


	events = [event.EventSwitchEnter, event.EventSwitchLeave,
			event.EventSwitchReconnected,
			event.EventPortAdd, event.EventPortDelete,
			event.EventPortModify,
			event.EventLinkAdd, event.EventLinkDelete]

	#发送控制信息
	@set_ev_cls(ofp_event.EventOFPSwitchFeatures, CONFIG_DISPATCHER)
	def send_controll_msg(self, ev):
		msg = ev.msg
		datapath = msg.datapath
		ofp = datapath.ofproto
		ofp_parser = datapath.ofproto_parser
		print('ffff', self.choice)
		# add table-miss
		match = ofp_parser.OFPMatch()
		actions = [ofp_parser.OFPActionOutput(ofp.OFPP_CONTROLLER, ofp.OFPCML_NO_BUFFER)]
		self.add_flows(datapath=datapath, priority=0, match=match, actions=actions)

	def add_flows(self, datapath, priority, match, actions):
		ofp = datapath.ofproto
		ofp_parser = datapath.ofproto_parser
		command = ofp.OFPFC_ADD
		inst = [ofp_parser.OFPInstructionActions(ofp.OFPIT_APPLY_ACTIONS, actions)]
		req = ofp_parser.OFPFlowMod(datapath=datapath, command=command,
									priority=priority, match=match, instructions=inst)
		print('Loading....')
		datapath.send_msg(req)
		print('Success!!!')
	
	@set_ev_cls(ofp_event.EventOFPStateChange, [MAIN_DISPATCHER, DEAD_DISPATCHER])
	#通过ryu.controller.handler.set_ev_cls装饰器（decorator）进行注册，在运行时，ryu控制器就能知道Monitor这个模块的函数_state_change_handler监听了一个事件
	def _state_change_handler(self, event):  #交换机状态发生变化后，让控制器数据于交换机一致
		datapath = event.datapath
		if event.state == MAIN_DISPATCHER:   # 在MAIN_DISPATCHER状态下，交换机处于上线状态
			if datapath.id not in self.datapaths:
				self.logger.debug('register datapath: %016x', datapath.id)    #十六进制输出
				self.datapaths[datapath.id] = datapath
			elif event.state == DEAD_DISPATCHER:
				if datapath.id in self.datapaths:
					self.logger.debug('unregister datapath %016x', datapath.id)
					del self.datapaths[datapath.id]

	#对交换机发送请求,获取终端信息
	def _monitor(self):
		while True:			#对已注册交换机发出统计信息获取请求每5秒无限地重复一次
			for dp in self.datapaths.values():
				self._request_stats(dp)
				self.monitor_info.setdefault(dp.id, [])
			self.send_echo_request()
			self.link_delay()
			self.update_topo()
			hub.sleep(2)

	def _request_stats(self, datapath):
		self.logger.debug('send stats request: %016x', datapath.id)
		#定义openflow解析器
		ofproto = datapath.ofproto
		ofp_parser = datapath.ofproto_parser

		#发送流量状态请求，获取流量状态数据
		request = ofp_parser.OFPFlowStatsRequest(datapath)
		datapath.send_msg(request)
		#
		request = ofp_parser.OFPPortStatsRequest(datapath, 0, ofproto.OFPP_ANY)
		datapath.send_msg(request)

	@set_ev_cls(ofp_event.EventOFPPortStatsReply, MAIN_DISPATCHER)
	def _port_stats_reply_handler(self, event):
		body = event.msg.body   #消息体
		self.logger.info('datapath         port      '
                         'rx-pkts  rx-bytes rx-error '  
                         'tx-pkts  tx-bytes tx-error ')    # rx-pkts:receive packets tx-pks:transmit packets
		self.logger.info('---------------- -------- '
						'-------- -------- -------- '
						'-------- -------- -------- ')
		
		tmp_key = ('port', 'rx-pkts', 'rx-bytes', 'rx-error', 'tx-pkts', 'tx-bytes', 'tx-error', 'rx-speed', 'tx-speed')
		tmp_dict = dict.fromkeys(tmp_key)
		tmp_list = []
		#print(str(self.worksheet.cell(self.worksheet.max_row,1).value))
		if str(self.worksheet.cell(self.worksheet.max_row,1).value) != self.NowDate and self.count == 0:
			print('fff', self.worksheet.max_row)
			for i in range(5):
				if i+1 == 1:
					self.worksheet.cell(self.worksheet.max_row+1,i+1).value = self.NowDate
					continue
					#print(self.worksheet.cell(self.worksheet.max_row,i+1).value)
				self.worksheet.cell(self.worksheet.max_row,i+1).value = 0
				#print(self.worksheet.cell(self.worksheet.max_row,i+1).value)
		for stat in sorted(body,key=attrgetter('port_no')):     #attrgetter：属性获取工具
			self.logger.info('%016x %8x %8d %8d %8d %8d %8d %8d',
							event.msg.datapath.id, stat.port_no,
							stat.rx_packets, stat.rx_bytes, stat.rx_errors,
							stat.tx_packets, stat.tx_bytes, stat.tx_errors)

			#存入monitor_info，提供给前段进行查询
			tmp_dict['port'] = stat.port_no
			tmp_dict['rx-pkts'] = stat.rx_packets
			tmp_dict['rx-bytes'] = stat.rx_bytes
			tmp_dict['rx-error'] = stat.rx_errors
			tmp_dict['tx-pkts'] = stat.tx_packets
			tmp_dict['tx-bytes'] = stat.tx_bytes
			tmp_dict['tx-error'] = stat.tx_errors
			self.rx_prebyte = self.rx_curbyte
			self.rx_curbyte = stat.rx_bytes
			tmp_dict['rx-speed'] = ((self.rx_curbyte - self.rx_prebyte) / 2)/1024
			self.tx_prebyte = self.tx_curbyte
			self.tx_curbyte = stat.tx_bytes
			tmp_dict['tx-speed'] = ((self.tx_curbyte - self.tx_prebyte) / 2)/1024
			tmp_list.append(tmp_dict.copy())
			self.monitor_info[event.msg.datapath.id] = tmp_list

			self.count += 1
			if str(self.worksheet.cell(self.worksheet.max_row,1).value) == self.NowDate:
				print(self.count)
				if stat.rx_bytes > 0:
					tmp11 = stat.rx_bytes/1024
					print('tmp1', stat.rx_bytes)
					self.worksheet.cell(self.worksheet.max_row,2).value = round(tmp11,2)
				if stat.tx_bytes > 0:
					tmp2 = stat.tx_bytes/1024
					print('tmp2', stat.tx_bytes)
					self.worksheet.cell(self.worksheet.max_row,3).value = round(tmp2,2)
				if self.count > 1:
					tmp3 = (self.worksheet.cell(self.worksheet.max_row,4).value + stat.rx_packets) / 2
					self.worksheet.cell(self.worksheet.max_row,4).value = round(tmp3,2)
					tmp4 = (self.worksheet.cell(self.worksheet.max_row,5).value + stat.tx_packets) / 2
					self.worksheet.cell(self.worksheet.max_row,5).value = round(tmp4,2)
				else:
					tmp3 = (self.worksheet.cell(self.worksheet.max_row,4).value + stat.rx_packets)
					self.worksheet.cell(self.worksheet.max_row,4).value = round(tmp3,2)
					tmp4 = (self.worksheet.cell(self.worksheet.max_row,5).value + stat.tx_packets)
					self.worksheet.cell(self.worksheet.max_row,5).value = round(tmp4,2)
			self.workbook.save("DateInfo.xlsx")	
			

	@set_ev_cls(ofp_event.EventOFPFlowStatsReply, MAIN_DISPATCHER)
	def _flow_stats_reply_handler(self, event):
		body = event.msg.body
		self.logger.info('datapath         '
						'in-port  eth-dst           '
						'out-port packets  bytes')
		self.logger.info('---------------- '
						'-------- ----------------- '
						'-------- -------- --------')
		for stat in sorted([flow for flow in body if flow.priority==1]
							,key=lambda flow:(flow.match['in_port'],flow.match['eth_dst'])):
			self.logger.info('%016x %8x %17s %8x %8d %8d',
							event.msg.datapath.id,stat.match['in_port'],
							stat.match['eth_dst'],stat.instructions[0].actions[0].port,
 							stat.packet_count,stat.byte_count)


	# 获取网络链路拓扑
    # 即将节点和边信息写入有向图中，默认的权重为0
	@set_ev_cls(events)
	def get_topo(self, event):
		switch_list = get_switch(self.topology_api_app)
		topo_switches = []
		# 得到每个设备的id，并写入图中作为图的节点
		for switch in switch_list:
			topo_switches.append(switch.dp.id)
		self.G.add_nodes_from(topo_switches)

		link_list = get_link(self.topology_api_app)
		self.links_src_dst = []
		# 将得到的链路的信息作为边写入图中
        # 注意这里links_src_dst时列表里列表，即[[],[],[]]，不能是元组，因为元组不可更改，也就是后面无法更新权重信息
		for link in link_list:
			self.links_src_dst.append([link.src.dpid, link.dst.dpid, 0])
		self.G.add_weighted_edges_from(self.links_src_dst)

		for link in link_list:
			self.links_src_dst.append([link.dst.dpid, link.src.dpid, 0])
		self.G.add_weighted_edges_from(self.links_src_dst)

	# 更新拓扑信息，主要更新有向图的边的权重
    # 即，程序获取链路的实时时延，当时延变化时，就将新的时延作为权重写入有向图中
	def update_topo(self):
    	# [[1, 2, 0], [3, 2, 0], [2, 1, 0], [2, 3, 0], [2, 1, 0], [2, 3, 0], [1, 2, 0], [3, 2, 0]]
        # {'2-3-3': 0.000362396240234375, '2-2-1': 0.001207113265991211, '1-2-2': 0.0004553794860839844, '3-2-2': 0.00015854835510253906}
        # 将link_Delay的时延作为权重更新进links_src_dst列表中，然后更新入有向图
		for key in self.link_Delay:
			list = key.split('_')
			l = (int(list[0]), int(list[2]))
			for i in self.links_src_dst:
				if l == (i[0], i[1]):
					i[2] = self.link_Delay[key]

		self.G.add_weighted_edges_from(self.links_src_dst)

	# 获取输出的端口，这里的输出端口是控制器指示数据转发时按照最短权重获得的输出端口进行数据转发
	def get_out_port(self, datapath, src, dst, in_port):
		global out_port
		dpid = datapath.id

		# 开始时，各个主机可能在图中不存在，因为开始ryu只获取了交换机的dpid，并不知道各主机的信息，
		# 所以需要将主机存入图中
		# 同时将记录主机和交换机之间的连接关系和端口
		if src not in self.G:
			self.G.add_node(src)
			self.G.add_weighted_edges_from([[dpid, src, 0]])
			self.G.add_weighted_edges_from([[src, dpid, 0]])
			src_dst = "%s-%s" % (dpid, src)
			self.id_port[src_dst] = in_port
			self.paths.setdefault(src, {})

		# 计算出基于最小权重的链路，按照这个转发链路进行数据的转发
		if dst in self.G and self.choice == 1:
			path = nx.shortest_path(self.G, src, dst, weight='weight')
			next_hop = path[path.index(dpid) + 1]
			for key in self.id_port:
				match_key = "%s-%s" % (dpid, next_hop)
				if key == match_key:
					out_port = self.id_port[key]
					# print('key_out_port:', out_port)
			print(path)
		#寻找最小跳数转发
		elif dst in self.G and self.choice == 2:
			if dst not in self.paths[src]:
				path = nx.shortest_path(self.G, src, dst) # 使用dijkstra算法
				self.paths[src][dst] = path
			path = self.paths[src][dst]
			next_hop = path[path.index(dpid) + 1] # 即把path列表里存的下一跳取出来，并不是做计算
			out_port = self.network[dpid][next_hop]['port']
			print(path)
		else:
			out_port = datapath.ofproto.OFPP_FLOOD
		return out_port

	# 由控制器向交换机发送echo报文，同时记录此时时间
	def send_echo_request(self):
		# 循环遍历交换机，逐一向存在的交换机发送echo探测报文
		for datapath in self.dpidSwitch.values():
			parser = datapath.ofproto_parser
			echo_req = parser.OFPEchoRequest(datapath, data=bytes("%.12f" % time.time(), encoding="utf8"))  # 获取当前时间

			datapath.send_msg(echo_req)
			# 每隔0.5秒向下一个交换机发送echo报文，防止回送报文同时到达控制器
			hub.sleep(0.5)
	
	# 交换机向控制器的echo请求回应报文，收到此报文时，控制器通过当前时间-时间戳，计算出往返时延
	@set_ev_cls(ofp_event.EventOFPEchoReply, [MAIN_DISPATCHER, CONFIG_DISPATCHER, HANDSHAKE_DISPATCHER])
	def echo_reply_handler(self, ev):
		now_timestamp = time.time()
		try:
			echo_delay = now_timestamp - eval(ev.msg.data)
			# 将交换机对应的echo时延写入字典保存起来
			self.echoDelay[ev.msg.datapath.id] = echo_delay
		except Exception as error:
			print("echo error:", error)
			return

	# 处理由交换机到来的消息，如LLDP消息和数据转发的消息
	@set_ev_cls(ofp_event.EventOFPPacketIn, MAIN_DISPATCHER)
	def packet_in_handler(self, ev):
		msg = ev.msg
		datapath = msg.datapath
		ofp = datapath.ofproto
		ofp_parser = datapath.ofproto_parser
		dpid = datapath.id
		in_port = msg.match['in_port']

		pkt = packet.Packet(msg.data)
		eth = pkt.get_protocols(ethernet.ethernet)[0]

		dst = eth.dst
		src = eth.src

        # try...except，由于packetin中存在LLDP消息和主机的数据转发消息，
        # 二者格式不一样，所以用try...except进行控制，分别处理两种消息；
		try:  # 处理到达的LLDP报文，从而获得LLDP时延
			src_dpid, src_outport = LLDPPacket.lldp_parse(msg.data)  # 获取两个相邻交换机的源交换机dpid和port_no(与目的交换机相连的端口)
			dst_dpid = msg.datapath.id  # 获取目的交换机（第二个），因为来到控制器的消息是由第二个（目的）交换机上传过来的
			if self.switches is None:
				self.switches = lookup_service_brick("switches")  # 获取交换机模块实例

			# 获得key（Port类实例）和data（PortData类实例）
			for port in self.switches.ports.keys():  # 开始获取对应交换机端口的发送时间戳
				if src_dpid == port.dpid and src_outport == port.port_no:  # 匹配key
					port_data = self.switches.ports[port]  # 获取满足key条件的values值PortData实例，内部保存了发送LLDP报文时的timestamp信息
					timestamp = port_data.timestamp
					if timestamp:
						delay = time.time() - timestamp
						self._save_delay_data(src=src_dpid, dst=dst_dpid, src_port=src_outport, lldp_dealy=delay)
		except Exception as error:  # 处理到达的主机的转发消息
			out_port = self.get_out_port(datapath, src, dst, in_port)
			actions = [ofp_parser.OFPActionOutput(out_port)]

            # 这里如果使用add_flow()进行了流表的添加，那么程序中的实时更新拓扑的权重就无意义了，转发就会依据流表进行
            # 所以这里不使用add_flow()方法，而是采用hub的形式，也就是每次转发都会请求控制器进行实时计算链路质量

            # 如果执行的动作不是flood，那么此时应该依据流表项进行转发操作，所以需要添加流表到交换机
            # if out_port != ofp.OFPP_FLOOD:
            #     match = ofp_parser.OFPMatch(in_port=in_port, eth_dst=dst, eth_src=src)
            #     self.add_flow(datapath=datapath, priority=1, match=match, actions=actions)

			data = None
			if msg.buffer_id == ofp.OFP_NO_BUFFER:
				data = msg.data
			# 控制器指导执行的命令
			out = ofp_parser.OFPPacketOut(datapath=datapath, buffer_id=msg.buffer_id,
											in_port=in_port, actions=actions, data=data)

	# 用于存储各个LLDP的时延
    # 同时记录拓扑中各个交换机之间的连接
	def _save_delay_data(self, src, dst, src_port, lldp_dealy):
		key = "%s-%s-%s" % (src, src_port, dst)
		src_dst = "%s-%s" % (src, dst)
		self.id_port[src_dst] = src_port
		# {'1-2': 2, '3-2': 2, '2-1': 2, '2-3': 3}
		self.src_dstDelay[key] = lldp_dealy

	# 计算链路的时延
	def link_delay(self):
		for key in self.src_dstDelay:
			list = key.split('-')
			t1 = 0
			t2 = 0
			for key_s in self.echoDelay:
				if key_s == int(list[0]):
					t1 = self.echoDelay[key_s]
				if key_s == int(list[2]):
					t2 = self.echoDelay[key_s]
			delay = self.src_dstDelay[key] - (t1 + t2) / 2
			# 由于误差及其他原因，可能出现时延为负值情况，如果为负值就不要进行时延的更新
			if delay >= 0:
				self.link_Delay[key] = self.src_dstDelay[key] - (t1 + t2) / 2
			else:
				continue



class Monitor_Controller(ControllerBase):

	def __init__(self, req, link, data, **config):
		super(Monitor_Controller, self).__init__(req, link, data, **config)
		self.monitor_app = data[monitor_instance_name]
		self.workbook = openpyxl.load_workbook("DateInfo.xlsx")
		self.worksheet = self.workbook.get_sheet_by_name('Sheet1')

	@route('monitor', urlQuery, methods=['GET'])
	def get_monitor_info(self, req, **kwargs):
		monitor_name = self.monitor_app

		if not monitor_name.monitor_info:
			return Response(status=404)
		body = json.dumps(monitor_name.monitor_info)
		return Response(content_type='application/json', body=body)
	
	@route('monitor', urlTotalData, methods=['GET'])
	def get_Total_Data(self, req, **kwargs):
		data = []
		tmp1 = []
		tmp2 = [] 	#输出流量(Kb)
		tmp3 = [] 	#输入流量(Kb)
		tmp4 = []  	#输出平均速率(pkts)
		tmp5 = []	#输入平均速率(pkts)
		index = self.worksheet.max_row
		print('yyy',index)
		for i in range(7):
			if index-i == 1:
				break
			tmp1.append(self.worksheet.cell(index-i,1).value)
			tmp2.append(self.worksheet.cell(index-i,2).value)
			tmp3.append(self.worksheet.cell(index-i,3).value)
			tmp4.append(self.worksheet.cell(index-i,4).value)
			tmp5.append(self.worksheet.cell(index-i,5).value)
		data.append(tmp1)
		data.append(tmp2)
		data.append(tmp3)
		data.append(tmp4)
		data.append(tmp5)
		body = json.dumps(data, ensure_ascii=False)
		return Response(content_type='application/json', body=body)

	@route('monitor', urlSetDtc, methods=['POST'])
	def set_detector(self, req, **kwargs):
		cid = kwargs['cid']
		self.choice = cid
		print(self.choice)
		print('reserved!!!')
		return Response(status=200)
